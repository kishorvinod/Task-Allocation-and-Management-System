from datetime import datetime, timedelta
from pathlib import Path
from typing import List, Dict, Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.database import db


class AnalyticsService:

    @staticmethod
    def workload_summary() -> List[Dict[str, Any]]:
        users = list(
            db.users.find({
                "role": {"$ne": "admin"}
            })
        )
        result = []

        for user in users:
            assigned_tasks = list(
                db.tasks.find({"assignedUser": user["_id"]})
            )

            allocated_hours = sum(
                task.get("estimatedHours", 0) for task in assigned_tasks
            )

            available_hours = (
                user.get("availableHoursPerDay", 0)
                * len(user.get("workingDays", []))
            )

            result.append({
                "userId": str(user["_id"]),
                "name": user.get("name"),
                "availableHours": available_hours,
                "allocatedHours": allocated_hours,
                "remainingHours": available_hours - allocated_hours,
                "taskCount": len(assigned_tasks),
            })

        return result

    @staticmethod
    def overloaded_users() -> List[Dict[str, Any]]:
        users = AnalyticsService.workload_summary()

        return [
            user for user in users
            if user["allocatedHours"] > user["availableHours"]
        ]

    @staticmethod
    def underutilized_users() -> List[Dict[str, Any]]:
        users = AnalyticsService.workload_summary()
        result = []

        for user in users:
            available = user["availableHours"]
            allocated = user["allocatedHours"]

            if available == 0:
                continue

            utilization = (allocated / available) * 100

            if utilization < 50:
                user["utilization"] = round(utilization, 2)
                result.append(user)

        return result

    @staticmethod
    def tasks_at_risk() -> List[Dict[str, Any]]:
        today = datetime.utcnow()
        threshold = today + timedelta(days=3)

        tasks = list(
            db.tasks.find({
                "status": {"$ne": "Completed"},
                "dueDate": {"$lte": threshold}
            })
        )

        result = []

        for task in tasks:
            due_date = task.get("dueDate")

            days_remaining = (
                (due_date - today).days if due_date else None
            )

            result.append({
                "taskId": str(task["_id"]),
                "title": task.get("title"),
                "status": task.get("status"),
                "dueDate": due_date,
                "daysRemaining": days_remaining,
            })

        return result

    @staticmethod
    def skill_demand() -> List[Dict[str, Any]]:
        tasks = list(
            db.tasks.aggregate([
                {
                    "$group": {
                        "_id": "$requiredSkill",
                        "tasks": {"$sum": 1}
                    }
                }
            ])
        )

        result = []

        for task in tasks:
            skill = task["_id"]

            user_count = db.users.count_documents({
                "role": {"$ne": "admin"},
                "skills": skill
            })

            result.append({
                "skill": skill,
                "users": user_count,
                "tasks": task["tasks"],
            })

        return result

    @staticmethod
    def task_completion_summary() -> Dict[str, Any]:
        total = db.tasks.count_documents({})
        completed = db.tasks.count_documents({"status": "Completed"})
        pending = db.tasks.count_documents({"status": "Pending"})
        in_progress = db.tasks.count_documents({"status": "In Progress"})

        completion_rate = (completed / total * 100) if total > 0 else 0

        return {
            "totalTasks": total,
            "completedTasks": completed,
            "pendingTasks": pending,
            "inProgressTasks": in_progress,
            "completionRate": round(completion_rate, 2),
        }

    @staticmethod
    def task_completion_summary_excel() -> Path:
        summary = AnalyticsService.task_completion_summary()
        data_dir = Path(__file__).resolve().parents[1] / "data"
        data_dir.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.utcnow().strftime("%Y%m%d%H%M%S")
        file_path = (
            data_dir /
            f"task-completion-summary-{timestamp}.xlsx"
        )

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Completion Summary"

        worksheet.append([
            "Metric",
            "Value"
        ])
        worksheet.append([
            "Total Tasks",
            summary["totalTasks"]
        ])
        worksheet.append([
            "Pending Tasks",
            summary["pendingTasks"]
        ])
        worksheet.append([
            "In Progress Tasks",
            summary["inProgressTasks"]
        ])
        worksheet.append([
            "Completed Tasks",
            summary["completedTasks"]
        ])
        worksheet.append([
            "Completion Rate",
            f"{summary['completionRate']}%"
        ])
        worksheet.append([
            "Generated At UTC",
            datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
        ])

        header_fill = PatternFill(
            fill_type="solid",
            fgColor="DBEAFE"
        )

        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill

        worksheet.column_dimensions["A"].width = 26
        worksheet.column_dimensions["B"].width = 22

        workbook.save(file_path)

        return file_path

    @staticmethod
    def priority_workload() -> List[Dict[str, Any]]:
        return list(
            db.tasks.aggregate([
                {
                    "$group": {
                        "_id": "$priority",
                        "tasks": {"$sum": 1},
                        "estimatedHours": {"$sum": "$estimatedHours"},
                    }
                }
            ])
        )

    @staticmethod
    def estimated_hours_by_status() -> List[Dict[str, Any]]:
        return list(
            db.tasks.aggregate([
                {
                    "$group": {
                        "_id": "$status",
                        "hours": {"$sum": "$estimatedHours"},
                    }
                }
            ])
        )
